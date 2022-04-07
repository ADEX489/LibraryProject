import pandas as pd
import numpy
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import date

def LibriPossesso():
    #Aprire documento
    cartella = os.getcwd()
    os.chdir(cartella)

    LibriInPossesso = {}

    LibriInPossessoP = pd.read_excel(getPath(),  sheet_name='LibriInPossesso')
    righe = LibriInPossessoP.shape[0]
    numero = 0

    while numero < righe:
        LibriInPossesso[LibriInPossessoP["ID PRESTITO"][numero]] = {"ID Libro": LibriInPossessoP["ID LIBRO"][numero], "ID Utente": LibriInPossessoP["ID UTENTE"][numero], "Data presa": LibriInPossessoP["DATA PRESA"][numero], "Data restituzione": LibriInPossessoP["DATA RESTITUZIONE"][numero]}
        numero += 1

    return LibriInPossesso

def BiblioP():
    # Aprire documento
    cartella = os.getcwd()
    os.chdir(cartella)

    LibriInPossessoP = pd.read_excel(getPath())

    return LibriInPossessoP

def printaS(spazio, dizionario):
    if spazio == "TUTTO":
        print(BiblioP())
    else:
        for index, item in dizionario.items():
            print(index, item[spazio])

def LibriXNome(IDUtente, dizionario):
    Libri = {}
    for x in dizionario:
        if(dizionario[x]['ID Utente'] == IDUtente ):
            Libri[IDUtente] = x
    return Libri

def X_Y(dizionario):
    righe = 0
    colonne = 0
    for x in dizionario:
        righe += 1
    for y in dizionario[x]:
        colonne += 1
    return [righe, colonne]

def CodiceXNome(codice, dizionario):
    for x in dizionario:
        if x == codice:
            return dizionario[x]['Nome']

def NumeroItems(dizionario):
    numero = 0
    for x in dizionario:
        numero += 1
    return numero


def CodiceXTitolo(codice, dizionario):
    for x in dizionario:
        if x == codice:
            return dizionario[x]['Titolo']

def CodiceLXCodice(Codice, dizionario):
    for x in dizionario:
        if dizionario[x]['ID Libro'] == Codice:
            return dizionario[x]['ID Utente']

def GetStato(codice, dizionario):
    for indexX, item in dizionario.items():
        if item['ID Libro'] == codice:
            return "si"

def getPath():
    return "FilesOperations\\a.xlsx"

def add(idLibro, idUtente, idPrestito):
    wb = load_workbook(getPath())
    ws = wb['LibriInPossesso']
    a = LibriPossesso()
    leng = len(a)

    ws.cell(column=1, row=leng + 2, value=idLibro)
    ws.cell(column=2, row=leng + 2, value=idUtente)
    ws.cell(column=3, row=leng + 2, value=date.today().strftime('%Y-%m-%d'))
    ws.cell(column=4, row=leng + 2, value="")
    ws.cell(column=5, row=leng + 2, value=idPrestito)
    wb.save(getPath())
    wb.close()

def DizionarioGirato(dizionario):
    ListaGirata = {}

    colonne = getColumns(dizionario)

    ListaGirata = {colonne[0].upper(): [], colonne[1].upper(): [], colonne[2].upper(): [], colonne[3].upper(): [],
                   "ID PRESTITO": []}
    for colonna in colonne:
        for codice, item in dizionario.items():
            ListaGirata[colonna.upper()].append(item[colonna])
    for x in dizionario:
        ListaGirata["ID PRESTITO"].append(x)

    return ListaGirata

def getColumns(dizionario):
    colonne = []
    for x in dizionario:
        c = x
    for x in dizionario[c]:
        colonne.append(x)
    return colonne

def writeIntoFile(data):
    data = pd.DataFrame(data)

    data.to_excel(getPath(), sheet_name='LibriInPossesso')