from openpyxl import Workbook, load_workbook
# If you need to get the column letter, also import this
from openpyxl.utils import get_column_letter

wb = load_workbook('C:\\Users\\ADEX_\\PycharmProjects\\pythonProject1\\FilesOperations\\a2.xlsx')
for sheet in wb:
    print(sheet)

ws = wb['LibriInPossesso']

print(ws['B2'].value)

append = ["ciao", "a", "tutti", ":)"]
ws.append(append)
wb.save('C:\\Users\\ADEX_\\PycharmProjects\\pythonProject1\\FilesOperations\\a2.xlsx')